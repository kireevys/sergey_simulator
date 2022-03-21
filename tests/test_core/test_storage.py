from datetime import datetime
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.core import ClosureAct, Order, OrderStatus
from app.storages import ExcelStorage
from tests.test_core.conftest import config

storage_root = config.storage["ROOT"]


class TestExcelStorage:
    def test_add_order(self, root):
        order = Order(
            order_id=12747295,
            warehouse_id=4734,
            description="ПРИКЛЕИТЬ СТИКЕР НА ВИТРИНУ",
            date=datetime.strptime("08.09.2021", "%d.%m.%Y"),
            status=OrderStatus.NEW,
        )
        path = Path(root, "2021/2021_учет_заявок.xlsx")

        storage = ExcelStorage(config)
        storage.add_order(order)

        wb = load_workbook(path, read_only=True)
        ws: Worksheet = wb.get_sheet_by_name("9")

        assert path.is_file()
        assert len(wb.worksheets) == 2
        assert ws.max_row == 2
        assert path.stat().st_size > 8000

        assert ws.cell(2, 1).value == order.order_id
        assert datetime.strptime(ws.cell(2, 2).value, "%d.%m.%Y") == order.date
        assert ws.cell(2, 3).value == order.warehouse_id
        assert ws.cell(2, 6).value == order.description

        assert storage._index.get(str(order.order_id)) == f"{path}:{order.date.month}:2"

    @pytest.mark.parametrize(
        "order,path",
        [
            (
                Order(
                    order_id=12747295,
                    warehouse_id=4734,
                    description="ПРИКЛЕИТЬ СТИКЕР НА ВИТРИНУ",
                    date=datetime.strptime("08.09.2021", "%d.%m.%Y"),
                    status=OrderStatus.NEW,
                ),
                Path(storage_root, "2021/2021_учет_заявок.xlsx"),
            ),
            (
                Order(
                    order_id=12747295,
                    warehouse_id=4734,
                    description="ПРИКЛЕИТЬ СТИКЕР НА ВИТРИНУ",
                    date=datetime.strptime("08.09.2022", "%d.%m.%Y"),
                    status=OrderStatus.NEW,
                ),
                Path(storage_root, "2022/2022_учет_заявок.xlsx"),
            ),
        ],
    )
    def test__get_wb_name_if_not_exists(self, order, path):
        storage = ExcelStorage(config)

        result = storage._get_wb_path(order)

        assert result == path

    @pytest.mark.parametrize(
        "order,path",
        [
            (
                Order(
                    order_id=12747295,
                    warehouse_id=4734,
                    description="ПРИКЛЕИТЬ СТИКЕР НА ВИТРИНУ",
                    date=datetime.strptime("08.09.2021", "%d.%m.%Y"),
                    status=OrderStatus.NEW,
                ),
                Path(storage_root, "2021/2021_учет_заявок.xlsx"),
            ),
            (
                Order(
                    order_id=12747295,
                    warehouse_id=4734,
                    description="ПРИКЛЕИТЬ СТИКЕР НА ВИТРИНУ",
                    date=datetime.strptime("08.09.2022", "%d.%m.%Y"),
                    status=OrderStatus.NEW,
                ),
                Path(storage_root, "2022/2022_учет_заявок.xlsx"),
            ),
        ],
    )
    def test__get_wb_by_order(self, order, path):
        storage = ExcelStorage(config)

        result = storage._get_wb_by_order(order)

        assert path.is_file()
        assert isinstance(result, Workbook)

    @pytest.mark.parametrize(
        "order,path,sheet_name",
        [
            (
                Order(
                    order_id=12747295,
                    warehouse_id=4734,
                    description="ПРИКЛЕИТЬ СТИКЕР НА ВИТРИНУ",
                    date=datetime.strptime("08.09.2021", "%d.%m.%Y"),
                    status=OrderStatus.NEW,
                ),
                Path(storage_root, "2021/2021_учет_заявок.xlsx"),
                "9",
            ),
            (
                Order(
                    order_id=12747295,
                    warehouse_id=4734,
                    description="ПРИКЛЕИТЬ СТИКЕР НА ВИТРИНУ",
                    date=datetime.strptime("08.12.2022", "%d.%m.%Y"),
                    status=OrderStatus.NEW,
                ),
                Path(storage_root, "2022/2022_учет_заявок.xlsx"),
                "12",
            ),
        ],
    )
    def test__get_ws_by_order(self, order, path, sheet_name):
        storage = ExcelStorage(config)
        path.parent.mkdir(parents=True, exist_ok=True)
        expected_wb = Workbook()
        expected_wb.create_sheet(sheet_name)

        result = storage._get_ws_by_order(order, expected_wb)

        assert result.title == sheet_name
        assert isinstance(result, Worksheet)

    @pytest.mark.parametrize(
        "order",
        [
            Order(
                order_id=1274725,
                warehouse_id=4734,
                description="ПРИКЛЕИТЬ СТИКЕР НА ВИТРИНУ",
                date=datetime.strptime("08.09.2021", "%d.%m.%Y"),
                status=OrderStatus.NEW,
            ),
            Order(
                order_id=12742725,
                warehouse_id=4734,
                description="ПРИКЛЕИТЬ",
                date=datetime.strptime("01.09.2021", "%d.%m.%Y"),
                status=OrderStatus.NEW,
            ),
        ],
    )
    def test_get_order_by_id(self, order):
        storage = ExcelStorage(config)
        storage.add_order(order)

        result = storage.get_order_by_id(order.order_id)

        assert result == order

    def test_close_order(self, root):
        date = datetime.strptime("08.09.2021", "%d.%m.%Y")
        order = Order(
            order_id=12747295,
            warehouse_id=4734,
            description="ПРИКЛЕИТЬ СТИКЕР НА ВИТРИНУ",
            date=datetime.strptime("08.09.2021", "%d.%m.%Y"),
            status=OrderStatus.NEW,
        )
        path = Path(root, "2021/2021_учет_заявок.xlsx")

        storage = ExcelStorage(config)
        storage.add_order(order)

        act = ClosureAct(order_id=order.order_id, date=date)
        storage.close_order(act)

        wb = load_workbook(path, read_only=True)
        ws: Worksheet = wb.get_sheet_by_name("9")

        assert path.is_file()
        assert len(wb.worksheets) == 2
        assert ws.max_row == 2
        assert path.stat().st_size > 8000

        assert ws.cell(2, 1).value == order.order_id
        assert datetime.strptime(ws.cell(2, 2).value, "%d.%m.%Y") == order.date
        assert ws.cell(2, 3).value == order.warehouse_id
        assert ws.cell(2, 6).value == order.description

        assert ws.cell(2, 9).value == act.date.strftime("%d.%m.%Y")
        assert ws.cell(2, 10).value == "Да"
