from datetime import datetime
from pathlib import Path

import pytest
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.core import Order, OrderStatus
from app.main import process
from app.storages import ExcelStorage
from tests.test_core.conftest import config

storage_root = config.storage["ROOT"]


@pytest.mark.parametrize(
    "email,order",
    [
        (
            Path("data/email.eml"),
            Order(
                order_id=12747295,
                warehouse_id=4734,
                description="ПРИКЛЕИТЬ СТИКЕР НА ВИТРИНУ",
                date=datetime.strptime("08.09.2021", "%d.%m.%Y"),
                status=OrderStatus.NEW,
            ),
        ),
        (
            Path("data/eng_email.eml"),
            Order(
                order_id=12767279,
                warehouse_id=5456,
                description="ТРЕБУЕТСЯ ЗАМЕНИТЬ ЛАМПЫ НА СКЛАДЕ",
                date=datetime.strptime("20.09.2021", "%d.%m.%Y"),
                status=OrderStatus.NEW,
            ),
        ),
    ],
)
def test_process(root, email, order):
    path = Path(root, "2021/2021_учет_заявок.xlsx")

    process(email, ExcelStorage(config))

    wb = load_workbook(path, read_only=True)
    ws: Worksheet = wb.get_sheet_by_name("9")

    assert path.is_file()
    assert len(wb.worksheets) == 2
    assert ws.max_row == 2
    assert path.stat().st_size > 8000

    assert ws.cell(2, 1).value == order.order_id
    assert datetime.strptime(ws.cell(2, 2).value, "%d.%m.%Y") == order.date
    assert ws.cell(2, 3).value == order.warehouse_id
    assert ws.cell(2, 6).value == order.description
