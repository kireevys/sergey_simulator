import asyncio
import logging
import time
from asyncio import get_event_loop
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.config import Config
from app.core import Order, OrderParser, Storage
from app.index import FileIndex
from app.parsers import EmailCAPParser, EmailOrderParser
from app.storages import ExcelStorage

logger = logging.getLogger()


def handle_exception(loop, context):
    logger.error("handled")


async def get_order(eml_path: Path, parser: OrderParser) -> (Order, Path):
    try:
        order = parser.parse(eml_path)
    except Exception:
        logger.error(f"{eml_path} parse error")
        raise

    logger.info(f"Parsed: {order.order_id}")
    return order, eml_path


def add_to_storage(order: Order, eml_path: Path, storage: Storage) -> bool:
    if not storage.is_exists(order):
        storage.add_order(order)
        storage.add_attachment(order, eml_path)
        logger.info(f"Added: {order.order_id}")
        return True
    else:
        logger.warning(f"{order.order_id} already exists")
        return False


def process(eml_path: Path, storage: Storage) -> Order:
    order = EmailOrderParser().parse(eml_path)
    add_to_storage(order, eml_path, storage)

    return order


def run(config: str, email: str):
    config = Config(Path(config))
    storage = ExcelStorage(config)

    process(Path(email), storage)


def close_orders(closure_path: str, config: str):
    closure_path = Path(closure_path)
    config = Config(Path(config))
    storage = ExcelStorage(config)
    parser = EmailCAPParser()

    for i in closure_path.glob("*.eml"):
        closure_act = parser.parse(i)
        try:
            order = storage.get_order_by_id(closure_act.order_id)
        except KeyError:
            logger.warning(f"Order not found: {closure_act.order_id}")
            continue

        storage.close_order(closure_act)
        storage.add_attachment(order, i)
        for a in closure_path.glob(f"{order.order_id}.*"):
            storage.add_attachment(order, a)


def bulk(config: str, emails: str):
    dir = Path(emails)
    config = Config(Path(config))
    storage = ExcelStorage(config)
    parser = EmailOrderParser()

    loop = get_event_loop()
    loop.set_exception_handler(handle_exception)

    tasks = [loop.create_task(get_order(email, parser)) for email in dir.rglob("*.eml")]

    total = len(tasks)
    logger.info(f"Start by {total} emails")

    start = time.time()
    results = loop.run_until_complete(asyncio.gather(*tasks, return_exceptions=True))

    logger.info(f"Parsed on {round(time.time() - start, 2)} seconds")

    logger.info(f"Writing to storage")
    start = time.time()
    for n, i in enumerate(results):

        if isinstance(i, Exception):
            continue

        order, eml_path = i

        add_to_storage(order, eml_path, storage)
        logger.info(f"Progress: {n:>5}/{total}")

    logger.info(f"Wrote on {round(time.time() - start, 2)} seconds")


def fill_index(config: str):
    config = Config(Path(config))
    storage = ExcelStorage(config)

    for i in Path(config.storage["ROOT"]).iterdir():
        if not i.is_dir():
            continue
        xlsx = i / f"{i.name}_учет_заявок.xlsx"
        wb = load_workbook(xlsx)
        for s in wb.sheetnames:
            if s == "default":
                continue
            ws: Worksheet = wb[s]
            for row in ws.iter_rows(min_row=2):
                index = FileIndex(config)
                index.add(storage.build_from_row(row), row[0].row)
