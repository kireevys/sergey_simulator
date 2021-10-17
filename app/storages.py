import logging
import shutil
from datetime import datetime
from functools import cached_property
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.config import Config, Template
from app.core import ClosureAct, Order, OrderStatus, Storage
from app.index import FileIndex

logger = logging.getLogger()


class ExcelStorage(Storage):
    def __init__(self, config: Config):
        self.config = config

    def _get_cwd(self, order: Order) -> Path:
        return Path(self.config.storage["ROOT"], str(order.date.year))

    def _get_wb_path(self, order: Order) -> Path:
        path = self._get_cwd(order) / f"{order.date.year}_учет_заявок.xlsx"
        if not path.is_file():
            path.parent.mkdir(exist_ok=True, parents=True)

        return path

    def _get_sheet_name(self, order: Order):
        return str(order.date.month)

    def _create_by_template(self, order: Order):
        path = self._get_wb_path(order)
        template = Template(self.config).get_template_by_name("orders_book")
        shutil.copy(template, path)

    def _get_wb_by_order(self, order: Order) -> Workbook:
        path = self._get_wb_path(order)

        if not path.is_file():
            self._create_by_template(order)

        return load_workbook(filename=path)

    def _get_ws_by_order(self, order: Order, wb: Workbook) -> Worksheet:
        name = self._get_sheet_name(order)
        try:
            ws = wb.get_sheet_by_name(name)
        except KeyError:
            tmpl = wb.get_sheet_by_name("default")
            ws = wb.copy_worksheet(tmpl)
            ws.title = name

        return ws

    def is_exists(self, order: Order) -> bool:
        return Path(
            self._get_cwd(order), order.date.strftime("%m"), str(order.order_id)
        ).is_dir()

    def add_order(self, order: Order):
        wb = self._get_wb_by_order(order)
        ws = self._get_ws_by_order(order, wb)
        order_as_row = [
            order.order_id,
            order.date.strftime("%d.%m.%Y"),
            order.warehouse_id,
            None,
            None,
            order.description,
        ]
        ws.append(order_as_row)
        logger.info(f"{order.order_id} added")
        wb.save(self._get_wb_path(order))

        self._index.add(order, ws.max_row)

    @cached_property
    def _index(self):
        return FileIndex(self.config)

    def build_from_row(self, row: tuple) -> Order:
        return Order(
            order_id=row[0].value,
            date=datetime.strptime(row[1].value, "%d.%m.%Y"),
            warehouse_id=row[2].value,
            description=row[5].value,
            # TODO: Возможно, придется когда то поправить
            status=OrderStatus.NEW,
        )

    def get_order_by_id(self, order_id: int) -> Order:
        path, sheet, row = self._index.get(str(order_id)).split(":")
        wb = load_workbook(filename=path)

        ws: Worksheet = wb.get_sheet_by_name(sheet)

        row = int(row)
        ws_order_id = ws.cell(row, 1).value
        assert ws_order_id == order_id
        rw = next(ws.iter_rows(min_row=row, max_row=row))
        return self.build_from_row(rw)

    def add_attachment(self, order: Order, attachment: Path):
        path = self._get_cwd(order)
        cwd = Path(path, order.date.strftime("%m"), str(order.order_id))
        cwd.mkdir(parents=True, exist_ok=True)
        shutil.copy(attachment, cwd)

    def close_order(self, act: ClosureAct):
        path, sheet, row = self._index.get(str(act.order_id)).split(":")
        row = int(row)
        wb = load_workbook(filename=path)

        ws = wb.get_sheet_by_name(sheet)

        ws.cell(row, 9).value = act.date.strftime("%d.%m.%Y")
        ws.cell(row, 10).value = "Да"

        wb.save(path)

        logger.info(
            f"Order {act.order_id} has been closed {act.date.strftime('%d.%m.%Y')}"
        )
